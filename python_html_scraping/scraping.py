import os
import sys
import unicodedata

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def get_html(url):
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"サイトへのアクセスに失敗しました:{e}")
        return None
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    return soup


def get_elements(soup):
    quotes = soup.find_all("div", class_="quote")
    if not quotes:
        return None

    quote_list = []
    author_list = []
    tag_list = []

    for quote in quotes:
        quote_text = quote.find("span", class_="text").text
        author_text = quote.find("small", class_="author").text
        tag_text = ",".join(tag.text for tag in quote.find_all("a", class_="tag"))
        quote_list.append(quote_text)
        author_list.append(author_text)
        tag_list.append(tag_text)
    return quote_list, author_list, tag_list


def overwrite_save(file_name, new_df):
    if os.path.exists(file_name):
        old_df = pd.read_excel(file_name)

        df = pd.concat([old_df, new_df], ignore_index=True)
        df = df.drop_duplicates(subset="quote", keep="first")

    else:
        df = new_df

    try:
        df.to_excel(file_name, index=False)
        return file_name
    except PermissionError as e:
        print("ファイルが開かれている可能性があります。閉じてから再実行してください")
        return None


def edit_excel(file_name):
    """excelファイルを編集する

    Args:
        file_name (str): 編集したいexcelファイルの名前
    """
    wb = load_workbook(file_name)
    ws = wb.active
    for cell in ws["A1:C1"][0]:
        cell.font = Font(bold=True)

    columns_list = ["A", "B", "C"]

    for column_name in columns_list:
        width_max = 0
        for cell in ws[column_name]:
            width_after = get_display_width(cell.value)
            width_max = max(width_max, width_after)
        ws.column_dimensions[column_name].width = width_max + 2
    try:
        wb.save(file_name)
        print("無事編集完了しました")
        return True
    except PermissionError as e:
        print("ファイルが開かれている可能性があります。閉じてから再実行してください")
        return None


def get_display_width(text):
    if text is None:
        return 0
    width = 0
    for word in text:
        if unicodedata.east_asian_width(word) in ("F", "W", "A"):
            width += 2
        else:
            width += 1
    return width


def create_dataframe(quote_list, author_list, tag_list):
    quote_dictionary = []
    for quote_num in range(len(quote_list)):
        quote_data = {
            "quote": quote_list[quote_num],
            "author": author_list[quote_num],
            "tags": tag_list[quote_num],
        }
        quote_dictionary.append(quote_data)
    df = pd.DataFrame(quote_dictionary)
    return df


if __name__ == "__main__":
    url = "https://quotes.toscrape.com/"
    soup = get_html(url)
    if soup is None:
        print("処理を終了します")
        sys.exit()
    result = get_elements(soup)
    if result is None:
        print("サイト情報を取得できなかったので、処理を終了します")
        sys.exit()
    quote_list, author_list, tag_list = result
    df = create_dataframe(quote_list, author_list, tag_list)
    file_name = os.path.join(BASE_DIR, "quote_list.xlsx")
    file_name = overwrite_save(file_name, df)
    if file_name is None:
        print("処理を終了します")
        sys.exit()
    edit_file = edit_excel(file_name)
    if edit_file is None:
        print("処理を終了します")
        sys.exit()
