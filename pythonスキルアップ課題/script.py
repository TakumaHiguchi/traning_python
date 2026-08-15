import feedparser
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import sys
import unicodedata


RSS_URL = "https://news.yahoo.co.jp/rss/topics/it.xml"


def create_dataframe(URL):
    """Yahoo newsのRSSにアクセスして、タイトル・URL・日時を取得し、
       データフレームを返す

    Args:
        URL (str): yahoo news RSS のURL
    """
    try:
        feed = feedparser.parse(URL)

        if feed.get("bozo", 0) == 1:
            print("データの取得に失敗したか、URLが正しくありません")
        else:
            news_list = []
            if not feed.entries:
                return None
            for entry in feed.entries:
                news_list.append(
                    {
                        "タイトル": entry.title,
                        "URL": entry.link,
                        "日付": entry.published,
                    }
                )

            df = pd.DataFrame(news_list)
            print("DataFrameの作成に成功しました")
            return df

    except Exception as e:
        print(f"エラーが発生しました:{e}")
        return None


def edit_excel(file_name):
    """excelファイルを編集する

    Args:
        file_name (str): 編集したいexcelファイルの名前
    """
    wb = load_workbook(file_name)
    ws = wb.active
    for cell in ws["A1:C1"][0]:
        cell.font = Font(bold=True)

    columns_list = ["A", "B", "C"]

    for column_name in columns_list:
        width_max = 0
        for cell in ws[column_name]:
            width_after = get_display_width(cell.value)
            width_max = max(width_max, width_after)
        ws.column_dimensions[column_name].width = width_max + 2
    try:
        wb.save(file_name)
        print("無事編集完了しました")
        return True
    except PermissionError as e:
        print("ファイルが開かれている可能性があります。閉じてから再実行してください")
        return None


def overwrite_save(file_name, new_df):
    if os.path.exists(file_name):
        old_df = pd.read_excel(file_name)

        df = pd.concat([old_df, new_df], ignore_index=True)
        df = df.drop_duplicates(subset="URL", keep="first")

    else:
        df = new_df

    try:
        df.to_excel(file_name, index=False)
        return file_name
    except PermissionError as e:
        print("ファイルが開かれている可能性があります。閉じてから再実行してください")
        return None


def get_display_width(text):
    width = 0
    for word in text:
        if unicodedata.east_asian_width(word) in ("F", "W", "A"):
            width += 2
        else:
            width += 1
    return width


if __name__ == "__main__":
    new_df = create_dataframe(RSS_URL)
    if new_df is None:
        print("処理を終了します")
        sys.exit()
    file_name = overwrite_save("news_list.xlsx", new_df)
    if file_name is None:
        print("処理を終了します")
        sys.exit()
    edit_file = edit_excel(file_name)
    if edit_file is None:
        print("処理を終了します")
        sys.exit()
